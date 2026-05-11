from __future__ import annotations

import argparse
import json
import re
from collections import defaultdict
from datetime import UTC, date, datetime
from pathlib import Path
from typing import Any

import openpyxl

FORBIDDEN_PUBLIC_TEXT = re.compile(
    r"("
    r"SUPABASE|SERVICE_ROLE|service_role|secret|password|authorization|bearer|"
    r"api[_-]?key|apikey|token|client_secret|OPENDART|BUILDING_REGISTER|"
    r"NAVER|admin|관리자|review|audit|감사|검토|formula|수식|"
    r"source_payload|source_row_hash|script|deposit|보증금|"
    r"specialTerms|특약|특수\s*계약|계약서"
    r")",
    re.IGNORECASE,
)

FORBIDDEN_PUBLIC_KEYS = {
    "deposit",
    "sourceRow",
    "sourceRows",
    "rowValues",
    "email",
    "이메일 주소",
    "담당자",
    "임대보증금",
    "보험 관련 특수 계약 조건",
    "기타 각종 특수 계약 조건",
}


def clean(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, str):
        text = value.strip()
        return text if text else None
    return value


def rows_from_sheet(ws, header_row: int, start_row: int) -> tuple[list[str], list[dict[str, Any]]]:
    header_values = list(next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True)))
    headers: list[str] = []
    counts: dict[str, int] = {}
    for index, header in enumerate(header_values, start=1):
        label = clean(header)
        if not label:
            headers.append(f"col_{index}")
            continue
        key = str(label)
        counts[key] = counts.get(key, 0) + 1
        headers.append(key if counts[key] == 1 else f"{key}__{index}")

    records: list[dict[str, Any]] = []
    for row_number, row in enumerate(ws.iter_rows(min_row=start_row, values_only=True), start=start_row):
        values = [clean(value) for value in row]
        if not any(value is not None for value in values):
            continue
        record = {headers[i]: values[i] if i < len(values) else None for i in range(len(headers))}
        record["_rowNumber"] = row_number
        records.append(record)
    return headers, records


def num(value: Any) -> float:
    if value in (None, ""):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).replace(",", ""))
    except ValueError:
        return 0.0


def first(*values: Any) -> Any:
    for value in values:
        if value not in (None, ""):
            return value
    return None


def mask_business_number(value: Any) -> Any:
    if not value:
        return None
    text = str(value)
    digits = re.sub(r"\D", "", text)
    if len(digits) < 5:
        return "***"
    return f"{digits[:3]}-{digits[3:5]}-*****"


def public_manager_label(team: Any) -> str:
    team_text = clean(team)
    return f"{team_text} 담당자" if team_text else "담당자 비공개"


def public_issue_content(row: dict[str, Any]) -> str:
    sheet = clean(row.get("시트")) or "원본"
    checked = bool(row.get("확인여부"))
    return f"{sheet} 데이터 품질 {'확인 완료' if checked else '확인 필요'}"


def is_forbidden_public_value(value: Any) -> bool:
    return isinstance(value, str) and bool(FORBIDDEN_PUBLIC_TEXT.search(value))


def is_safe_field_dictionary_row(row: dict[str, Any]) -> bool:
    public_text = " ".join(str(value) for value in row.values() if value not in (None, ""))
    return not FORBIDDEN_PUBLIC_TEXT.search(public_text)


def assert_public_dashboard_safe(payload: dict[str, Any]) -> None:
    findings: list[str] = []

    def walk(value: Any, path: str = "") -> None:
        if isinstance(value, dict):
            for key, child in value.items():
                next_path = f"{path}.{key}" if path else str(key)
                if key in FORBIDDEN_PUBLIC_KEYS or FORBIDDEN_PUBLIC_TEXT.search(str(key)):
                    findings.append(next_path)
                walk(child, next_path)
        elif isinstance(value, list):
            for index, child in enumerate(value):
                walk(child, f"{path}[{index}]")
        elif is_forbidden_public_value(value):
            findings.append(path)

    walk(payload)
    if findings:
        sample = ", ".join(findings[:30])
        raise ValueError(f"Public dashboard payload contains forbidden fields/text: {sample}")


def asset_id(asset_code: Any) -> str:
    code = str(asset_code or "unknown").strip().lower()
    safe = "".join(ch if ch.isalnum() else "_" for ch in code)
    return f"asset_{safe}"


def tenant_id(name: Any, brn: Any) -> str:
    base = str(brn or name or "unknown").strip().lower()
    safe = "".join(ch if ch.isalnum() else "_" for ch in base)
    return f"tenant_{safe}"


def load_map_points(path: Path) -> dict[str, dict[str, Any]]:
    if not path.exists():
        return {}
    try:
        home = json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return {}
    points = {}
    for point in home.get("mapPoints", []):
        aid = point.get("assetId")
        if aid:
            points[aid] = {
                "assetId": aid,
                "latitude": point.get("latitude"),
                "longitude": point.get("longitude"),
                "address": point.get("address"),
            }
    return points


def build_dashboard(source_xlsx: Path, map_json: Path) -> dict[str, Any]:
    wb = openpyxl.load_workbook(source_xlsx, read_only=True, data_only=True)
    general_headers, general_rows = rows_from_sheet(wb["DB_일반"], 9, 12)
    history_headers, history_rows = rows_from_sheet(wb["DB_히스토리 누적"], 10, 15)
    _, manager_rows = rows_from_sheet(wb["자산_담당자 연결"], 3, 4)
    _, issue_rows = rows_from_sheet(wb["Log"], 4, 5)
    meta_headers, meta_rows = rows_from_sheet(wb["Meta_데이터 항목 설명"], 2, 3)
    map_points = load_map_points(map_json)

    assets: dict[str, dict[str, Any]] = {}
    tenants: dict[str, dict[str, Any]] = {}
    leases: list[dict[str, Any]] = []

    for row in general_rows:
        if row.get("펀드코드") in ("예시 데이터", "단위"):
            continue
        code = row.get("자산코드")
        name = row.get("자산명")
        tenant_name = row.get("임차인명")
        if not code or not name:
            continue
        aid = asset_id(code)
        tid = tenant_id(tenant_name, row.get("임차인 사업자번호"))
        leased_area = num(row.get("임대면적"))
        exclusive_area = num(row.get("전용면적"))
        gross_area = num(row.get("전체 연면적"))

        asset = assets.setdefault(
            aid,
            {
                "assetId": aid,
                "assetCode": code,
                "assetName": name,
                "fundCode": row.get("펀드코드"),
                "fundName": row.get("펀드명"),
                "sector": row.get("섹터"),
                "grossFloorAreaSqm": 0,
                "leasedAreaSqm": 0,
                "tenantIds": set(),
                "leaseCount": 0,
                "map": map_points.get(aid, {"assetId": aid}),
            },
        )
        asset["grossFloorAreaSqm"] = max(asset["grossFloorAreaSqm"], gross_area)
        asset["leasedAreaSqm"] += leased_area
        asset["tenantIds"].add(tid)
        asset["leaseCount"] += 1

        tenant = tenants.setdefault(
            tid,
            {
                "tenantId": tid,
                "tenantName": tenant_name or "-",
                "businessNumber": mask_business_number(row.get("임차인 사업자번호")),
                "assetIds": set(),
                "leasedAreaSqm": 0,
                "leaseCount": 0,
            },
        )
        tenant["assetIds"].add(aid)
        tenant["leasedAreaSqm"] += leased_area
        tenant["leaseCount"] += 1

        leases.append(
            {
                "leaseId": f"lease_{len(leases) + 1:04d}",
                "assetId": aid,
                "tenantId": tid,
                "assetName": name,
                "tenantName": tenant_name,
                "floor": row.get("임차 층"),
                "space": row.get("임차 세부 구역"),
                "goodsType": row.get("취급 상품 유형"),
                "coldStorage": row.get("저온창고 여부"),
                "preleased": row.get("선임차 여부"),
                "thirdPartyLogistics": row.get("3PL 여부"),
                "leasedAreaSqm": leased_area,
                "exclusiveAreaSqm": exclusive_area,
                "startDate": row.get("현재 계약개시일"),
                "endDate": row.get("현재 계약만기일"),
                "rf": row.get("RF"),
                "fo": row.get("FO"),
                "ti": row.get("TI"),
                "status": row.get("계약 상태"),
            }
        )

    rent_history: list[dict[str, Any]] = []
    for row in history_rows:
        if row.get("펀드코드") in ("예시 데이터", "단위"):
            continue
        if not row.get("자산코드") or not row.get("자산명"):
            continue
        rent_history.append(
            {
                "assetId": asset_id(row.get("자산코드")),
                "tenantId": tenant_id(row.get("임차인명"), row.get("임차인 사업자번호")),
                "assetName": row.get("자산명"),
                "tenantName": row.get("임차인명"),
                "floor": row.get("임차 층"),
                "space": row.get("임차 세부 구역"),
                "baseDate": row.get("기준일자"),
                "reason": row.get("임대료 변동 원인"),
                "monthlyRentTotal": num(row.get("월임대료 총액")),
                "monthlyMfTotal": num(row.get("월관리비 총액")),
                "rentPerPy": num(row.get("평당 월임대료")),
                "mfPerPy": num(row.get("평당 월관리비")),
                "leasedAreaSqm": num(row.get("임대면적")),
            }
        )

    rent_by_asset: dict[str, float] = defaultdict(float)
    rent_by_tenant: dict[str, float] = defaultdict(float)
    rent_by_month: dict[str, dict[str, float]] = defaultdict(lambda: {"monthlyRentTotal": 0.0, "monthlyMfTotal": 0.0})
    for row in rent_history:
        combined = row["monthlyRentTotal"] + row["monthlyMfTotal"]
        rent_by_asset[row["assetId"]] += combined
        rent_by_tenant[row["tenantId"]] += combined
        month = str(row.get("baseDate") or "")[:7] or "미기재"
        rent_by_month[month]["monthlyRentTotal"] += row["monthlyRentTotal"]
        rent_by_month[month]["monthlyMfTotal"] += row["monthlyMfTotal"]

    asset_list = []
    for asset in assets.values():
        asset["tenantCount"] = len(asset["tenantIds"])
        asset["tenantIds"] = sorted(asset["tenantIds"])
        asset["monthlyCostTotal"] = rent_by_asset.get(asset["assetId"], 0.0)
        asset["vacancyAreaSqm"] = max(asset["grossFloorAreaSqm"] - asset["leasedAreaSqm"], 0.0)
        asset["vacancyRate"] = asset["vacancyAreaSqm"] / asset["grossFloorAreaSqm"] if asset["grossFloorAreaSqm"] else 0
        asset_list.append(asset)

    tenant_list = []
    for tenant in tenants.values():
        tenant["assetCount"] = len(tenant["assetIds"])
        tenant["assetIds"] = sorted(tenant["assetIds"])
        tenant["monthlyCostTotal"] = rent_by_tenant.get(tenant["tenantId"], 0.0)
        tenant_list.append(tenant)

    managers = []
    for row in manager_rows:
        if not row.get("자산코드"):
            continue
        managers.append(
            {
                "assetId": asset_id(row.get("자산코드")),
                "assetCode": row.get("자산코드"),
                "assetName": row.get("자산명"),
                "manager": public_manager_label(row.get("소속")),
                "team": row.get("소속"),
            }
        )

    issues = []
    for row in issue_rows:
        if not row.get("순번"):
            continue
        issues.append(
            {
                "issueId": f"issue_r{int(num(row.get('_rowNumber'))):04d}_{int(num(row.get('순번'))):04d}" if num(row.get("순번")) else f"issue_r{int(num(row.get('_rowNumber'))):04d}_{len(issues)+1:04d}",
                "date": row.get("일자"),
                "checked": bool(row.get("확인여부")),
                "sheet": row.get("시트"),
                "asset": row.get("자산"),
                "content": public_issue_content(row),
                "severity": "확인 완료" if row.get("확인여부") else "확인 필요",
            }
        )

    monthly_trend = [
        {"month": month, **values, "monthlyCostTotal": values["monthlyRentTotal"] + values["monthlyMfTotal"]}
        for month, values in sorted(rent_by_month.items())
    ]

    summary = {
        "assetCount": len(asset_list),
        "tenantCount": len(tenant_list),
        "leaseCount": len(leases),
        "grossFloorAreaSqm": sum(a["grossFloorAreaSqm"] for a in asset_list),
        "leasedAreaSqm": sum(a["leasedAreaSqm"] for a in asset_list),
        "vacancyAreaSqm": sum(a["vacancyAreaSqm"] for a in asset_list),
        "monthlyCostTotal": sum(a["monthlyCostTotal"] for a in asset_list),
        "issueCount": len([issue for issue in issues if not issue["checked"]]),
    }
    summary["vacancyRate"] = summary["vacancyAreaSqm"] / summary["grossFloorAreaSqm"] if summary["grossFloorAreaSqm"] else 0

    dashboard = {
        "schemaVersion": "logistics_leasing_dashboard_v1",
        "generatedAt": datetime.now(UTC).isoformat(timespec="seconds").replace("+00:00", "Z"),
        "source": {
            "type": "xlsx_public_snapshot",
            "fileName": source_xlsx.name,
            "note": "xlsx 기준 읽기 전용 공개 데이터입니다.",
        },
        "summary": summary,
        "assets": sorted(asset_list, key=lambda item: str(item["assetName"])),
        "tenants": sorted(tenant_list, key=lambda item: str(item["tenantName"])),
        "leases": leases,
        "rentHistory": rent_history,
        "monthlyTrend": monthly_trend,
        "managers": managers,
        "issues": issues,
        "fieldDictionary": [
            {k: v for k, v in row.items() if not k.startswith("_")}
            for row in meta_rows
            if any(v is not None for k, v in row.items() if not k.startswith("_"))
            and is_safe_field_dictionary_row({k: v for k, v in row.items() if not k.startswith("_")})
        ],
        "preservation": {
            "tabs": ["Weekly", "Home", "Asset", "Company", "Sector", "Analysis Tools", "Data Playground", "Data Quality", "Permissions"],
            "requiredInteractions": ["kpi-modal", "table-row-drawer", "map-marker-drawer", "filter-refresh", "select-refresh", "chart-source-modal"],
        },
    }
    assert_public_dashboard_safe(dashboard)
    return dashboard


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source-xlsx", required=True)
    parser.add_argument("--map-json", required=True)
    parser.add_argument("--output", required=True)
    args = parser.parse_args()

    output = Path(args.output)
    output.parent.mkdir(parents=True, exist_ok=True)
    dashboard = build_dashboard(Path(args.source_xlsx), Path(args.map_json))
    output.write_text(json.dumps(dashboard, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({"output": str(output), "assets": len(dashboard["assets"]), "tenants": len(dashboard["tenants"]), "leases": len(dashboard["leases"])}, ensure_ascii=False))


if __name__ == "__main__":
    main()
